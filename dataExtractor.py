import pandas as pd
from flask import make_response
from io import BytesIO
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# List of month names
month_names = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]


def singleCategoryFormatter(values, formatted_district_data, catName):
    """
    Formats the data for a single category (e.g., index values, mothers values).

    Args:
        values (list): List of dictionaries containing data for each year.
        formatted_district_data (dict): Dictionary to store formatted data.
        catName (str): Name of the category.

    Returns:
        dict: Formatted district data.

    Raises:
        Exception: If an error occurs during processing.
    """
    try:
        for value in values:
            year = value["year"]
            monthly_data = value["data"]

            formatted_index_value = {"year": year, "singleYearData": []}

            for i, val in enumerate(monthly_data):
                month = month_names[i] if i < len(month_names) else f"Month_{i + 1}"
                formatted_monthly_data = {month: val}
                formatted_index_value["singleYearData"].append(formatted_monthly_data)
            formatted_district_data[catName].append(formatted_index_value)
        return formatted_district_data

    except Exception as e:
        raise Exception(f"Error processing the file: {str(e)}")


def read_database(mongo):
    """
    Retrieves data from a MongoDB collection based on the provided type parameter.

    Args:
        mongo: MongoDB instance.

    Returns:
        list: List of formatted data.

    Raises:
        Exception: If an error occurs during processing.
    """
    try:
        # MongoDB collection for monthly anemia data
        collection = mongo.db.anemiaDataMonthly
        # Retrieve documents from the collection
        documents = list(collection.find({}, {"_id": 0}))

        output_data = []

        for state_data in documents:
            state_name = state_data["state"]
            districts_data = state_data["data"]

            formatted_state_data = []

            for district_data in districts_data:
                district_name = district_data["District"]
                index_values = district_data.get("Index Value", [])
                rank_values = district_data.get("Rank", [])
                children6_59_months_values = district_data.get(
                    "Children (6 - 59 months)", []
                )
                children6_9_years_values = district_data.get(
                    "Children (6 - 9 years)", []
                )
                adolescents_10_19_years_values = district_data.get(
                    "Adolescents (10 - 19 years)", []
                )
                pregnant_women_values = district_data.get("Pregnant Women", [])
                mothers_values = district_data.get("Mothers", [])

                formatted_district_data = {
                    "district": district_name,
                    "index_values": [],
                    "rank_values": [],
                    "children6_59_months_values": [],
                    "children6_9_years_values": [],
                    "adolescents_10_19_years_values": [],
                    "pregnant_women_values": [],
                    "mothers_values": [],
                }

                # Format Index Values
                formatted_district_data = singleCategoryFormatter(
                    index_values, formatted_district_data, "index_values"
                )

                # Format children months
                formatted_district_data = singleCategoryFormatter(
                    children6_59_months_values,
                    formatted_district_data,
                    "children6_59_months_values",
                )

                # Format children years
                formatted_district_data = singleCategoryFormatter(
                    children6_9_years_values,
                    formatted_district_data,
                    "children6_9_years_values",
                )

                # Format adolescents years
                formatted_district_data = singleCategoryFormatter(
                    adolescents_10_19_years_values,
                    formatted_district_data,
                    "adolescents_10_19_years_values",
                )

                # Format pregnant women
                formatted_district_data = singleCategoryFormatter(
                    pregnant_women_values,
                    formatted_district_data,
                    "pregnant_women_values",
                )

                # Format mothers
                formatted_district_data = singleCategoryFormatter(
                    mothers_values,
                    formatted_district_data,
                    "mothers_values",
                )

                # Format state rank
                formatted_district_data = singleCategoryFormatter(
                    rank_values,
                    formatted_district_data,
                    "rank_values",
                )

                formatted_state_data.append(formatted_district_data)
            output_data.append(
                {"state": state_name, "districtsData": formatted_state_data}
            )
        return output_data
    except Exception as e:
        raise Exception(f"Error processing the file: {str(e)}")


def modifyToExcel(data):
    """
    Modifies the provided data and exports it to an Excel file.

    Args:
        data (list): List of formatted data.

    Returns:
        flask.Response: Excel file as a Flask response.

    Raises:
        Exception: If an error occurs during processing.
    """
    try:
        state_dfs = []
        for state_data in data:
            state_name = state_data["state"]
            year_dfs = []

            for district_data in state_data["districtsData"]:
                district_name = district_data["district"]

                rows = []
                for (
                    index_value,
                    mothers_value,
                    children_months_value,
                    children_years_value,
                    adolescents_value,
                    pregnant_women_value,
                    rank_value,
                ) in zip(
                    district_data["index_values"],
                    district_data["mothers_values"],
                    district_data["children6_59_months_values"],
                    district_data["children6_9_years_values"],
                    district_data["adolescents_10_19_years_values"],
                    district_data["pregnant_women_values"],
                    district_data["rank_values"],
                ):
                    year = index_value["year"]
                    index_value_data = index_value["singleYearData"]
                    mothers_data = mothers_value["singleYearData"]
                    children_months_data = children_months_value["singleYearData"]
                    children_years_data = children_years_value["singleYearData"]
                    adolescents_data = adolescents_value["singleYearData"]
                    pregnant_women_data = pregnant_women_value["singleYearData"]
                    rank_data = rank_value["singleYearData"]

                    for i, month in enumerate(month_names):
                        index_value = (
                            index_value_data[i] if i < len(index_value_data) else None
                        )
                        mothers_value = (
                            mothers_data[i] if i < len(mothers_data) else None
                        )

                        children_month_value = (
                            children_months_data[i]
                            if i < len(children_months_data)
                            else None
                        )
                        children_year_value = (
                            children_years_data[i]
                            if i < len(children_years_data)
                            else None
                        )
                        adolescent_value = (
                            adolescents_data[i] if i < len(adolescents_data) else None
                        )
                        pregnant_woman_value = (
                            pregnant_women_data[i]
                            if i < len(pregnant_women_data)
                            else None
                        )
                        ranks_value = rank_data[i] if i < len(rank_data) else None

                        rows.append(
                            {
                                "Month": month,
                                "Year": year,
                                "District": district_name,
                                "Index Value": index_value[month]
                                if index_value
                                else None,
                                "Mothers": mothers_value[month]
                                if mothers_value
                                else None,
                                "Children (6 - 59 months)": children_month_value[month]
                                if children_month_value
                                else None,
                                "Children (6 - 9 years)": children_year_value[month]
                                if children_year_value
                                else None,
                                "Adolescents": adolescent_value[month]
                                if adolescent_value
                                else None,
                                "Pregnant Women": pregnant_woman_value[month]
                                if pregnant_woman_value
                                else None,
                                "Rank": ranks_value[month] if ranks_value else None,
                            }
                        )

                df = pd.DataFrame(rows)
                year_dfs.append(df)

            for df in year_dfs:
                df["State"] = state_name

            state_df = pd.concat(year_dfs, ignore_index=True)
            state_dfs.append(state_df)

        final_df = pd.concat(state_dfs, ignore_index=True)
        columns_to_keep = [
            "Month",
            "Year",
            "State",
            "District",
            "Index Value",
            "Mothers",
            "Children (6 - 59 months)",
            "Children (6 - 9 years)",
            "Adolescents",
            "Pregnant Women",
            "Rank",
        ]
        final_df = final_df[columns_to_keep]

        column_order = [
            "Month",
            "Year",
            "State",
            "District",
            "Rank",
            "Index Value",
            "Children (6 - 59 months)",
            "Children (6 - 9 years)",
            "Adolescents",
            "Pregnant Women",
            "Mothers",
        ]
        final_df = final_df.reindex(columns=column_order, fill_value=None)

        excel_output = BytesIO()

        with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Sheet1")

            worksheet = writer.sheets["Sheet1"]

            (max_row, max_col) = final_df.shape

            table_range = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

            table = Table(displayName="MyTable", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)

            for row_num in range(1, max_row + 1):
                for col_num in range(max_col):
                    cell_coord = openpyxl.utils.get_column_letter(col_num + 1) + str(
                        row_num + 1
                    )
                    worksheet[cell_coord].alignment = Alignment(horizontal="center")

            for col_num, value in enumerate(final_df.columns.values):
                column_letter = openpyxl.utils.get_column_letter(col_num + 1)
                max_length = max(
                    final_df[value].astype(str).apply(len).max(), len(value)
                )
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_output.seek(0)

        response = make_response(excel_output.read())
        response.headers["Content-Disposition"] = "attachment; filename=output.xlsx"
        response.headers[
            "Content-Type"
        ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return response

    except Exception as e:
        import traceback

        print("Error details", traceback.format_exc())
        raise Exception(f"Error in file: {str(e)}")
